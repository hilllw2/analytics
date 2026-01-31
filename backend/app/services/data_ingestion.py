"""
Data Ingestion Service
Handles file upload, parsing, type inference, and data loading.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Tuple, Any, BinaryIO
from datetime import datetime
import chardet
import io
import re
import warnings
from pathlib import Path

from app.core.config import settings
from app.core.session_manager import DatasetInfo


class DataIngestionService:
    """
    Service for ingesting data files with robust handling.
    Supports CSV, TSV, XLSX, XLS with auto-detection.
    """
    
    # Common date formats to try
    DATE_FORMATS = [
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%m/%d/%Y',
        '%Y/%m/%d',
        '%d-%m-%Y',
        '%m-%d-%Y',
        '%Y-%m-%d %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%m/%d/%Y %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%dT%H:%M:%SZ',
        '%Y-%m-%d %H:%M:%S.%f',
        '%b %d, %Y',
        '%B %d, %Y',
        '%d %b %Y',
        '%d %B %Y',
        '%Y-%m-%dT%H:%M:%S.%f',
        '%Y-%m-%dT%H:%M:%S.%fZ',
        '%m/%d/%y',
        '%d/%m/%y',
        '%Y%m%d',
        '%d-%b-%Y',
        '%d-%B-%Y',
    ]
    
    def __init__(self):
        pass
    
    async def detect_file_type(self, filename: str, content: bytes) -> str:
        """Detect file type from filename and content."""
        ext = Path(filename).suffix.lower()
        
        if ext in ['.xlsx', '.xls']:
            return 'excel'
        elif ext == '.tsv':
            return 'tsv'
        elif ext == '.csv':
            return 'csv'
        else:
            # Try to detect from content
            sample = content[:1000].decode('utf-8', errors='ignore')
            if '\t' in sample and ',' not in sample:
                return 'tsv'
            return 'csv'
    
    async def detect_encoding(self, content: bytes) -> str:
        """Detect file encoding using chardet."""
        result = chardet.detect(content[:10000])
        encoding = result.get('encoding', 'utf-8')
        confidence = result.get('confidence', 0)
        
        # Fallback for low confidence
        if confidence < 0.7:
            # Try common encodings
            for enc in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    content[:1000].decode(enc)
                    return enc
                except:
                    continue
        
        return encoding or 'utf-8'
    
    async def detect_delimiter(self, content: bytes, encoding: str) -> str:
        """Auto-detect CSV delimiter."""
        sample = content[:5000].decode(encoding, errors='ignore')
        lines = sample.split('\n')[:10]
        
        delimiters = [',', '\t', ';', '|']
        scores = {}
        
        for delim in delimiters:
            counts = [line.count(delim) for line in lines if line.strip()]
            if counts and len(set(counts)) == 1 and counts[0] > 0:
                # Consistent count across lines
                scores[delim] = counts[0]
        
        if scores:
            return max(scores, key=scores.get)
        
        return ','  # Default
    
    async def get_excel_sheets(self, content: bytes) -> List[Dict[str, Any]]:
        """Get list of sheets in an Excel file with preview info."""
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
        sheets = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Get dimensions
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0
            
            # Get header preview
            headers = []
            if row_count > 0:
                for cell in ws[1]:
                    headers.append(str(cell.value) if cell.value else f"Column_{len(headers)+1}")
            
            sheets.append({
                "name": sheet_name,
                "row_count": row_count,
                "column_count": col_count,
                "headers": headers[:20]  # First 20 headers
            })
        
        wb.close()
        return sheets
    
    async def load_file(
        self,
        content: bytes,
        filename: str,
        sheet_name: Optional[str] = None,
        sample_mode: bool = False,
        max_rows: Optional[int] = None
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Load a file into a DataFrame with metadata.
        
        Args:
            content: File content as bytes
            filename: Original filename
            sheet_name: For Excel files, which sheet to load
            sample_mode: If True, load only a sample for preview
            max_rows: Maximum rows to load (None for all)
        
        Returns:
            Tuple of (DataFrame, metadata dict)
        """
        file_type = await self.detect_file_type(filename, content)
        metadata = {
            "original_filename": filename,
            "file_type": file_type,
            "encoding": None,
            "delimiter": None,
            "sheet_name": sheet_name,
            "is_sampled": False,
            "full_row_count": None,
            "warnings": []
        }
        
        nrows = None
        if sample_mode:
            nrows = settings.PREVIEW_ROWS
            metadata["is_sampled"] = True
        elif max_rows:
            nrows = max_rows
        
        try:
            if file_type == 'excel':
                df = await self._load_excel(content, sheet_name, nrows, metadata)
            else:
                df = await self._load_csv(content, file_type, nrows, metadata)
            
            # Post-processing
            df = self._clean_column_names(df)
            df = await self._infer_and_convert_types(df, metadata)
            
            metadata["row_count"] = len(df)
            metadata["column_count"] = len(df.columns)
            
            return df, metadata
            
        except Exception as e:
            raise ValueError(f"Failed to load file: {str(e)}")
    
    async def _load_excel(
        self,
        content: bytes,
        sheet_name: Optional[str],
        nrows: Optional[int],
        metadata: Dict
    ) -> pd.DataFrame:
        """Load Excel file."""
        from io import BytesIO
        
        buffer = BytesIO(content)
        
        # Default to first sheet if not specified
        if sheet_name is None:
            xl = pd.ExcelFile(buffer)
            sheet_name = xl.sheet_names[0]
            metadata["sheet_name"] = sheet_name
        
        df = pd.read_excel(
            buffer,
            sheet_name=sheet_name,
            nrows=nrows,
            engine='openpyxl'
        )
        
        # Get full row count if sampling
        if nrows:
            buffer.seek(0)
            full_df = pd.read_excel(buffer, sheet_name=sheet_name, engine='openpyxl')
            metadata["full_row_count"] = len(full_df)
            del full_df
        
        return df
    
    async def _load_csv(
        self,
        content: bytes,
        file_type: str,
        nrows: Optional[int],
        metadata: Dict
    ) -> pd.DataFrame:
        """Load CSV/TSV file."""
        encoding = await self.detect_encoding(content)
        metadata["encoding"] = encoding
        
        if file_type == 'tsv':
            delimiter = '\t'
        else:
            delimiter = await self.detect_delimiter(content, encoding)
        metadata["delimiter"] = delimiter
        
        # Try to load with detected settings
        try:
            df = pd.read_csv(
                io.BytesIO(content),
                encoding=encoding,
                delimiter=delimiter,
                nrows=nrows,
                low_memory=False,
                on_bad_lines='warn'
            )
        except Exception as e:
            # Fallback with more lenient settings
            metadata["warnings"].append(f"Initial parse failed, using fallback: {str(e)}")
            df = pd.read_csv(
                io.BytesIO(content),
                encoding='latin-1',
                delimiter=delimiter,
                nrows=nrows,
                low_memory=False,
                on_bad_lines='skip'
            )
        
        # Get full row count if sampling
        if nrows:
            full_count = sum(1 for _ in io.BytesIO(content)) - 1  # -1 for header
            metadata["full_row_count"] = full_count
        
        return df
    
    def _clean_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize column names."""
        new_columns = []
        for i, col in enumerate(df.columns):
            if pd.isna(col) or str(col).strip() == '':
                new_col = f"Column_{i+1}"
            else:
                new_col = str(col).strip()
                # Remove excessive whitespace
                new_col = re.sub(r'\s+', ' ', new_col)
            new_columns.append(new_col)
        
        # Handle duplicates
        seen = {}
        final_columns = []
        for col in new_columns:
            if col in seen:
                seen[col] += 1
                final_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                final_columns.append(col)
        
        df.columns = final_columns
        return df
    
    async def _infer_and_convert_types(
        self,
        df: pd.DataFrame,
        metadata: Dict
    ) -> pd.DataFrame:
        """Infer and convert column types."""
        inferred_types = {}
        date_columns = []
        numeric_columns = []
        categorical_columns = []
        
        for col in df.columns:
            try:
                col_type, converted = await self._infer_column_type(df[col])
                
                if converted is not None:
                    try:
                        df[col] = converted
                    except Exception:
                        # If conversion fails, keep original
                        pass
                
                inferred_types[col] = col_type
                
                if col_type == 'datetime':
                    date_columns.append(col)
                elif col_type in ['integer', 'float']:
                    numeric_columns.append(col)
                elif col_type in ['string', 'category']:
                    categorical_columns.append(col)
            except Exception as e:
                # If type inference fails for a column, mark it as string
                inferred_types[col] = 'string'
                categorical_columns.append(col)
        
        metadata["inferred_types"] = inferred_types
        metadata["date_columns"] = date_columns
        metadata["numeric_columns"] = numeric_columns
        metadata["categorical_columns"] = categorical_columns
        
        return df
    
    async def _infer_column_type(
        self,
        series: pd.Series
    ) -> Tuple[str, Optional[pd.Series]]:
        """
        Infer the best type for a column and optionally convert it.
        
        Returns:
            Tuple of (type_name, converted_series or None)
        """
        # Already typed correctly
        if pd.api.types.is_datetime64_any_dtype(series):
            return 'datetime', None
        
        if pd.api.types.is_bool_dtype(series):
            return 'boolean', None
        
        if pd.api.types.is_integer_dtype(series):
            return 'integer', None
        
        if pd.api.types.is_float_dtype(series):
            return 'float', None
        
        # Need to infer from string/object column
        non_null = series.dropna()
        if len(non_null) == 0:
            return 'empty', None
        
        sample = non_null.head(1000)
        
        # Try boolean
        try:
            bool_values = {'true', 'false', 'yes', 'no', '1', '0', 't', 'f', 'y', 'n'}
            sample_str = sample.astype(str).str.lower().str.strip()
            if sample_str.isin(bool_values).all():
                def convert_bool(x):
                    if pd.isna(x):
                        return None
                    val = str(x).lower().strip()
                    if val in {'true', 'yes', '1', 't', 'y'}:
                        return True
                    if val in {'false', 'no', '0', 'f', 'n'}:
                        return False
                    return None
                converted = series.apply(convert_bool)
                return 'boolean', converted
        except Exception:
            pass
        
        # Try numeric
        try:
            numeric = pd.to_numeric(sample, errors='coerce')
            if numeric.notna().sum() / len(sample) > 0.9:  # 90% convertible
                full_numeric = pd.to_numeric(series, errors='coerce')
                # Check if values are integers (handle NaN properly)
                non_null_numeric = full_numeric.dropna()
                if len(non_null_numeric) > 0 and (non_null_numeric == non_null_numeric.astype(int)).all():
                    return 'integer', full_numeric
                return 'float', full_numeric
        except Exception:
            pass
        
        # Try datetime
        for fmt in self.DATE_FORMATS:
            try:
                parsed = pd.to_datetime(sample, format=fmt, errors='coerce')
                if parsed.notna().sum() / len(sample) > 0.8:  # 80% parseable
                    full_parsed = pd.to_datetime(series, format=fmt, errors='coerce')
                    return 'datetime', full_parsed
            except Exception:
                continue
        
        # Try generic datetime parsing (mixed formats / dateutil)
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', UserWarning)
                parsed = pd.to_datetime(sample, errors='coerce')
            if parsed.notna().sum() / len(sample) > 0.8:
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', UserWarning)
                    full_parsed = pd.to_datetime(series, errors='coerce')
                return 'datetime', full_parsed
        except Exception:
            pass
        
        # Check cardinality for categorical - but don't convert, just mark as category
        try:
            unique_ratio = series.nunique() / len(series) if len(series) > 0 else 1
            if unique_ratio < 0.5 and series.nunique() < settings.HIGH_CARDINALITY_THRESHOLD:
                # Don't convert to categorical dtype, just mark it as category type
                return 'category', None
        except Exception:
            pass
        
        return 'string', None
    
    def create_dataset_info(
        self,
        name: str,
        df: pd.DataFrame,
        metadata: Dict[str, Any]
    ) -> DatasetInfo:
        """Create a DatasetInfo object from DataFrame and metadata."""
        return DatasetInfo(
            name=name,
            original_filename=metadata.get("original_filename", name),
            df=df,
            inferred_types=metadata.get("inferred_types", {}),
            date_columns=metadata.get("date_columns", []),
            numeric_columns=metadata.get("numeric_columns", []),
            categorical_columns=metadata.get("categorical_columns", []),
            upload_time=datetime.now(),
            row_count=len(df),
            column_count=len(df.columns),
            is_sampled=metadata.get("is_sampled", False),
            full_row_count=metadata.get("full_row_count")
        )


# Singleton instance
data_ingestion = DataIngestionService()
